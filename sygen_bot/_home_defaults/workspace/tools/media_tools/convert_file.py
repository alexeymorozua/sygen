#!/usr/bin/env python3
"""Universal file converter.

Supported conversions:
  text/markdown → PDF
  DOCX → text / PDF
  XLSX → text / CSV
  HEIC/HEIF → JPG/PNG
  image resize (any format)

Usage:
    python convert_file.py --input file.md --to pdf
    python convert_file.py --input file.docx --to txt
    python convert_file.py --input file.docx --to pdf
    python convert_file.py --input file.xlsx --to csv
    python convert_file.py --input file.heic --to jpg
    python convert_file.py --input file.png --resize 800

Returns JSON: {"output": "/path/to/result", "format": "pdf", ...}
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import shutil
import sys
from pathlib import Path

_OUTPUT_DIR = Path(__file__).parent.parent.parent / "output_to_user"
_FONT_PATH = Path("/Library/Fonts/Arial Unicode.ttf")
_FALLBACK_FONT = Path("/System/Library/Fonts/Helvetica.ttc")


def _ensure_output_dir() -> Path:
    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return _OUTPUT_DIR


def _out_path(stem: str, ext: str) -> Path:
    dest = _ensure_output_dir() / f"{stem}.{ext}"
    n = 1
    while dest.exists():
        dest = _ensure_output_dir() / f"{stem}_{n}.{ext}"
        n += 1
    return dest


# ── Text/Markdown → PDF ─────────────────────────────────────────────

def _md_to_pdf(input_path: Path) -> dict:
    from fpdf import FPDF

    text = input_path.read_text(encoding="utf-8")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    font = _FONT_PATH if _FONT_PATH.exists() else _FALLBACK_FONT
    pdf.add_font("main", "", str(font))
    pdf.set_font("main", size=11)

    def _mc(h: float, txt: str) -> None:
        pdf.set_x(pdf.l_margin)
        w = pdf.w - pdf.l_margin - pdf.r_margin
        pdf.multi_cell(w, h, txt)

    for line in text.splitlines():
        stripped = line.strip()

        if stripped.startswith("# "):
            pdf.set_font("main", size=18)
            _mc(10, stripped[2:].strip())
            pdf.set_font("main", size=11)
        elif stripped.startswith("## "):
            pdf.set_font("main", size=15)
            _mc(9, stripped[3:].strip())
            pdf.set_font("main", size=11)
        elif stripped.startswith("### "):
            pdf.set_font("main", size=13)
            _mc(8, stripped[4:].strip())
            pdf.set_font("main", size=11)
        elif stripped.startswith("- ") or stripped.startswith("* "):
            _mc(6, f"  \u2022 {stripped[2:].strip()}")
        elif stripped == "":
            pdf.ln(4)
        else:
            clean = re.sub(r"\*\*(.*?)\*\*", r"\1", stripped)
            clean = re.sub(r"\*(.*?)\*", r"\1", clean)
            clean = re.sub(r"`(.*?)`", r"\1", clean)
            _mc(6, clean)

    dest = _out_path(input_path.stem, "pdf")
    pdf.output(str(dest))
    return {"output": str(dest), "format": "pdf", "pages": pdf.pages_count}


# ── DOCX → text / PDF ───────────────────────────────────────────────

def _docx_to_text(input_path: Path) -> str:
    from docx import Document
    doc = Document(str(input_path))
    return "\n".join(p.text for p in doc.paragraphs)


def _docx_to_txt(input_path: Path) -> dict:
    text = _docx_to_text(input_path)
    dest = _out_path(input_path.stem, "txt")
    dest.write_text(text, encoding="utf-8")
    return {"output": str(dest), "format": "txt", "chars": len(text)}


def _docx_to_pdf(input_path: Path) -> dict:
    text = _docx_to_text(input_path)
    tmp = _ensure_output_dir() / f"_tmp_{input_path.stem}.md"
    tmp.write_text(text, encoding="utf-8")
    try:
        result = _md_to_pdf(tmp)
    finally:
        tmp.unlink(missing_ok=True)
    return result


# ── XLSX → text / CSV ───────────────────────────────────────────────

def _xlsx_to_csv(input_path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active
    dest = _out_path(input_path.stem, "csv")
    with dest.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([c if c is not None else "" for c in row])
    wb.close()
    return {"output": str(dest), "format": "csv"}


def _xlsx_to_txt(input_path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active
    lines = []
    for row in ws.iter_rows(values_only=True):
        lines.append("\t".join(str(c) if c is not None else "" for c in row))
    wb.close()
    text = "\n".join(lines)
    dest = _out_path(input_path.stem, "txt")
    dest.write_text(text, encoding="utf-8")
    return {"output": str(dest), "format": "txt", "rows": len(lines)}


# ── HEIC/Image conversion (sips) ────────────────────────────────────

def _image_convert(input_path: Path, target_fmt: str) -> dict:
    fmt_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "tiff": "tiff", "bmp": "bmp"}
    sips_fmt = fmt_map.get(target_fmt.lower())
    if not sips_fmt:
        return {"error": f"Unsupported image format: {target_fmt}"}

    dest = _out_path(input_path.stem, "jpg" if sips_fmt == "jpeg" else target_fmt.lower())
    try:
        subprocess.run(
            ["sips", "-s", "format", sips_fmt, str(input_path), "--out", str(dest)],
            capture_output=True, timeout=30, check=True,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        return {"error": f"sips failed: {e}"}
    return {"output": str(dest), "format": target_fmt.lower()}


# ── Image resize ─────────────────────────────────────────────────────

def _image_resize(input_path: Path, max_side: int) -> dict:
    ext = input_path.suffix.lstrip(".").lower()
    if ext == "heic":
        ext = "jpg"
    dest = _out_path(f"{input_path.stem}_{max_side}px", ext)

    shutil.copy2(input_path, dest)
    try:
        subprocess.run(
            ["sips", "--resampleHeightWidthMax", str(max_side), str(dest)],
            capture_output=True, timeout=30, check=True,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as e:
        dest.unlink(missing_ok=True)
        return {"error": f"sips resize failed: {e}"}
    return {"output": str(dest), "format": ext, "max_side": max_side}


# ── Router ───────────────────────────────────────────────────────────

_CONVERSIONS: dict[tuple[str, str], callable] = {
    ("md", "pdf"): _md_to_pdf,
    ("txt", "pdf"): _md_to_pdf,
    ("docx", "txt"): _docx_to_txt,
    ("docx", "pdf"): _docx_to_pdf,
    ("xlsx", "csv"): _xlsx_to_csv,
    ("xlsx", "txt"): _xlsx_to_txt,
    ("heic", "jpg"): lambda p: _image_convert(p, "jpg"),
    ("heic", "png"): lambda p: _image_convert(p, "png"),
    ("heif", "jpg"): lambda p: _image_convert(p, "jpg"),
    ("png", "jpg"): lambda p: _image_convert(p, "jpg"),
    ("jpg", "png"): lambda p: _image_convert(p, "png"),
    ("webp", "jpg"): lambda p: _image_convert(p, "jpg"),
    ("webp", "png"): lambda p: _image_convert(p, "png"),
}


def convert(input_path: Path, to_fmt: str, resize: int | None = None) -> dict:
    if not input_path.exists():
        return {"error": f"File not found: {input_path}"}

    if resize:
        return _image_resize(input_path, resize)

    src_ext = input_path.suffix.lstrip(".").lower()
    to_fmt = to_fmt.lower()

    fn = _CONVERSIONS.get((src_ext, to_fmt))
    if not fn:
        return {"error": f"Conversion {src_ext} → {to_fmt} not supported. Supported: {list(_CONVERSIONS.keys())}"}

    return fn(input_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Universal file converter")
    parser.add_argument("--input", required=True, help="Input file path")
    parser.add_argument("--to", default="pdf", help="Target format (pdf, txt, csv, jpg, png)")
    parser.add_argument("--resize", type=int, help="Resize image to max N px side")
    args = parser.parse_args()

    result = convert(Path(args.input), args.to, args.resize)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if "error" in result:
        sys.exit(1)


if __name__ == "__main__":
    main()
